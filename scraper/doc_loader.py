# scraper/doc_loader.py
import os
import fitz  # PyMuPDF
from docx import Document
from openpyxl import load_workbook
from scraper.site_parser import categorize_section
from knowledge.db import add_knowledge, clear_knowledge_by_source

def load_pdf(path: str, min_length: int = 40) -> list[str]:
    doc = fitz.open(path)
    chunks = []
    for page in doc:
        for line in page.get_text("text").split("\n"):
            line = line.strip()
            if len(line) >= min_length:
                chunks.append(line)
    return chunks

def load_docx(path: str, min_length: int = 30) -> list[str]:
    doc = Document(path)
    return [p.text.strip() for p in doc.paragraphs if len(p.text.strip()) >= min_length]

def load_xlsx(path: str) -> list[str]:
    wb = load_workbook(path, read_only=True)
    chunks = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                line = " | ".join(cells)
                if len(line) > 10:
                    chunks.append(line)
    finally:
        wb.close()
    return chunks

LOADERS = {".pdf": load_pdf, ".docx": load_docx, ".xlsx": load_xlsx}

async def load_all_docs(docs_dir: str = "docs"):
    for filename in os.listdir(docs_dir):
        ext = os.path.splitext(filename)[1].lower()
        if ext not in LOADERS:
            continue
        path = os.path.join(docs_dir, filename)
        await clear_knowledge_by_source(filename)
        try:
            chunks = LOADERS[ext](path)
        except Exception as e:
            print(f"Не удалось загрузить {filename}: {e}")
            continue
        for chunk in chunks:
            await add_knowledge(filename, categorize_section(chunk), chunk)
